"""
app.py  –  Lenovo Elite Engineer Certification Program
=======================================================
Self-contained single-file Streamlit application.
All components are inlined – no external file dependencies.

Run:
    pip install -r requirements.txt
    streamlit run app.py
"""

# ─────────────────────────────────────────────────────────────────────────────
# IMPORTS
# ─────────────────────────────────────────────────────────────────────────────
from __future__ import annotations

import asyncio
import base64
import io
import json
import time
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

import pandas as pd
import streamlit as st
import streamlit.components.v1 as components
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
# PAGE CONFIG  (must be first Streamlit call)
# ─────────────────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Lenovo Elite Engineer Certification",
    page_icon="🔧",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# ─────────────────────────────────────────────────────────────────────────────
# CONSTANTS
# ─────────────────────────────────────────────────────────────────────────────
PASS_THRESHOLD = 90   # minimum technical score (%)
RESULTS_FILE   = Path("assessment_results.xlsx")

ALEX_INTRO = (
    "Hello and welcome. My name is Alex, and I will be your virtual interviewer today. "
    "Before we begin, I'd like to briefly explain how this interview will be conducted. "
    "This assessment is designed to evaluate your knowledge, experience, and problem-solving "
    "abilities through verbal responses. "
    "To ensure a fair and consistent evaluation for all participants, the session is monitored "
    "for interview integrity. "
    "During the interview, our system may observe indicators such as: "
    "Keyboard activity. Window or tab switching. Extended periods of looking away from the screen. "
    "Mobile phone usage. Multiple people appearing in the camera frame. "
    "Extended absence from the camera view. "
    "Please note that these indicators do not automatically imply misconduct. "
    "They are simply used to calculate an Interview Integrity Score. "
    "Your Integrity Score begins at 100 points. "
    "Keyboard activity during a response: minus 10 points. "
    "Repeated keyboard activity: minus 20 points. "
    "Window or application switching: minus 25 points. "
    "Looking away from the screen for more than 5 seconds: minus 10 points. "
    "Mobile phone detected: minus 30 points. "
    "Multiple faces detected: minus 30 points. "
    "Camera view unavailable for more than 10 seconds: minus 20 points. "
    "At the end of the interview, two independent scores will be generated: "
    "Technical Competency Score and Interview Integrity Score. "
    "Both scores will be reviewed together as part of the overall evaluation process. "
    "To achieve the best experience, please ensure: "
    "Your camera remains on throughout the interview. "
    "Your face remains clearly visible. "
    "You answer all questions verbally. "
    "Mobile phones and other reference materials are kept away. "
    "Browser searches, external assistance, and note-taking are avoided. "
    "If you are ready, we can begin. I wish you the very best for your interview today."
)

QUESTIONS: List[Dict] = [
    {
        "id": 1, "topic": "No Power – Intermittent Power Rail Failure",
        "question": (
            "A Lenovo ThinkPad shows no signs of power. The AC adapter LED remains ON. "
            "USB-C power meter shows 20V negotiation, but pressing the power button causes "
            "a brief 0.2A current spike before dropping to 0A. The power button LED never "
            "illuminates. What is the most likely fault?"
        ),
        "options": {
            "A": "Corrupt BIOS image",
            "B": "Shorted system board power rail after DC-in stage",
            "C": "Defective LCD panel",
            "D": "Missing operating system",
        },
        "answer": "B",
        "explanation": (
            "Voltage negotiation is successful, indicating adapter and charging IC "
            "communication are functioning. The immediate current collapse suggests a "
            "downstream short circuit on a power rail."
        ),
    },
    {
        "id": 2, "topic": "No POST – Memory Initialization Failure",
        "question": (
            "A system powers on, fan spins continuously, Caps Lock LED responds to key "
            "presses, but there is no Lenovo logo or video output. External display also "
            "shows no signal. Diagnostic logs indicate POST halts during memory training. "
            "What should be checked first?"
        ),
        "options": {
            "A": "LCD cable connection",
            "B": "SSD firmware version",
            "C": "Memory module compatibility and seating",
            "D": "WLAN card",
        },
        "answer": "C",
        "explanation": (
            "POST stopping during memory training typically points to RAM-related issues "
            "before video initialization occurs."
        ),
    },
    {
        "id": 3, "topic": "No Power – Embedded Controller Behavior",
        "question": (
            "A laptop previously experienced liquid damage. Upon adapter connection, the "
            "charging LED blinks 3 times and turns off. The power button is unresponsive. "
            "Which component is most likely preventing startup?"
        ),
        "options": {
            "A": "SSD controller",
            "B": "Embedded Controller (EC) detecting fault condition",
            "C": "LCD backlight fuse",
            "D": "TPM module",
        },
        "answer": "B",
        "explanation": (
            "Modern systems often rely on EC logic to block startup when critical faults "
            "or abnormal sensor readings are detected."
        ),
    },
    {
        "id": 4, "topic": "Dim Display – Backlight Circuit Diagnosis",
        "question": (
            "A notebook boots successfully. An image is visible under a flashlight, but "
            "the screen appears almost black. External monitor functions normally. "
            "Which component is MOST likely defective?"
        ),
        "options": {
            "A": "CPU",
            "B": "GPU firmware",
            "C": "LCD backlight circuit or panel backlight",
            "D": "BIOS region",
        },
        "answer": "C",
        "explanation": (
            "Visible image under a flashlight confirms the video signal exists; "
            "only illumination is missing."
        ),
    },
    {
        "id": 5, "topic": "Crisis BIOS Recovery",
        "question": (
            "During a BIOS update, AC power is accidentally disconnected. The laptop now "
            "powers on with a black screen and maximum fan speed. Keyboard backlight turns "
            "on briefly. What is the BEST recovery action?"
        ),
        "options": {
            "A": "Replace LCD panel",
            "B": "Reinstall operating system",
            "C": "Perform Crisis BIOS Recovery using approved recovery media",
            "D": "Replace SSD",
        },
        "answer": "C",
        "explanation": (
            "Symptoms are classic BIOS corruption indicators. Crisis recovery is "
            "designed specifically for this scenario."
        ),
    },
    {
        "id": 6, "topic": "BIOS Flash Failure Analysis",
        "question": (
            "A BIOS flash utility reports success. After reboot, the system powers on "
            "but repeatedly restarts every 15 seconds without displaying the Lenovo splash "
            "screen. What is the most probable cause?"
        ),
        "options": {
            "A": "Corrupted EC firmware accompanying BIOS update",
            "B": "Damaged keyboard membrane",
            "C": "Defective battery pack only",
            "D": "Windows driver conflict",
        },
        "answer": "A",
        "explanation": (
            "Modern BIOS packages often update EC firmware. Incomplete EC programming "
            "can create endless reboot cycles."
        ),
    },
    {
        "id": 7, "topic": "No POST vs No Power Identification",
        "question": (
            "A technician observes: Power LED ON, fan spinning, keyboard backlight active, "
            "no Lenovo logo, no video output. "
            "How should the issue be categorized?"
        ),
        "options": {
            "A": "No Power",
            "B": "No Boot",
            "C": "No POST",
            "D": "Dim Display",
        },
        "answer": "C",
        "explanation": (
            "The system is receiving power and partially initializing hardware but is "
            "failing before POST completion."
        ),
    },
    {
        "id": 8, "topic": "Dim Display – Backlight Fuse Troubleshooting",
        "question": (
            "After a display replacement, the image is visible but extremely dim. A "
            "known-good LCD panel exhibits identical behavior. "
            "What should be investigated next?"
        ),
        "options": {
            "A": "SSD SMART status",
            "B": "Backlight enable signal and LCD fuse on motherboard",
            "C": "BIOS password settings",
            "D": "Operating system display scaling",
        },
        "answer": "B",
        "explanation": (
            "Since two panels exhibit the same symptom, the motherboard backlight "
            "circuit becomes the primary suspect."
        ),
    },
    {
        "id": 9, "topic": "Crisis BIOS – Recovery Media Detection",
        "question": (
            "A technician initiates Crisis BIOS Recovery. The USB drive LED flashes "
            "continuously for 3 minutes before the system automatically shuts down. "
            "What does this MOST likely indicate?"
        ),
        "options": {
            "A": "Recovery process successfully read BIOS image and is reflashing firmware",
            "B": "USB port failure",
            "C": "LCD cable fault",
            "D": "Battery authentication failure",
        },
        "answer": "A",
        "explanation": (
            "Continuous USB activity during recovery generally indicates successful "
            "reading and writing of firmware data."
        ),
    },
    {
        "id": 10, "topic": "Advanced Board-Level Diagnosis",
        "question": (
            "A laptop presents as No Power. Measurements reveal: 20V present at USB-C "
            "input, 3.3V always rail present, 5V always rail present, power button signal "
            "reaches EC, EC does not issue S0 power-on sequence. "
            "What is the MOST likely root cause?"
        ),
        "options": {
            "A": "Defective LCD panel",
            "B": "Corrupted EC firmware or EC hardware failure",
            "C": "Missing bootloader in Windows",
            "D": "Defective SSD",
        },
        "answer": "B",
        "explanation": (
            "Since primary rails and power button communication are functioning, failure "
            "of the EC to initiate the power sequence indicates EC firmware or hardware issues."
        ),
    },
    {
        "id": 11, "topic": "Bonus Challenge – Expert Level",
        "question": (
            "BONUS: A system exhibits: Power LED ON, fan spins briefly then stops, USB "
            "recovery key not detected, SPI flash voltage normal, POST analyzer card "
            "remains at code 00. What is the MOST likely diagnosis?"
        ),
        "options": {
            "A": "Corrupt operating system",
            "B": "LCD backlight failure",
            "C": "CPU initialization failure or corrupted boot block region of BIOS",
            "D": "SSD not detected",
        },
        "answer": "C",
        "explanation": (
            "POST code 00 typically indicates the CPU never begins execution or the BIOS "
            "boot block is damaged before initialization can occur."
        ),
    },
]

# ─────────────────────────────────────────────────────────────────────────────
# INLINE PROCTORING + AVATAR COMPONENT HTML
# ─────────────────────────────────────────────────────────────────────────────
# This HTML is injected directly via components.html() – no file needed.
# Python fills in {QUIZ_MODE} and {ALLOWED_IDS} before rendering.
PROCTORING_HTML_TEMPLATE = """\
<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8"/>
<style>
*{{box-sizing:border-box;margin:0;padding:0;}}
html,body{{background:#0F0F0F;color:#E0E0E0;font-family:'Segoe UI',system-ui,sans-serif;
           font-size:12px;height:100%;overflow:hidden;}}

/* ── Layout ─────────────────────────────────── */
.root{{display:flex;flex-direction:column;gap:8px;padding:10px;height:100%;}}
.top{{display:flex;gap:8px;}}

/* ── Avatar panel ────────────────────────────── */
.avatar-panel{{
  flex:1;background:#141414;border:1.5px solid #1E1E1E;border-radius:8px;
  display:flex;flex-direction:column;align-items:center;justify-content:center;
  padding:8px;position:relative;overflow:hidden;min-height:170px;
}}
.avatar-bg{{
  position:absolute;inset:0;
  background:radial-gradient(ellipse at 50% 30%,#1a0000 0%,#0f0f0f 70%);
}}
.avatar-wrap{{position:relative;z-index:1;}}
.avatar-name{{position:relative;z-index:1;font-size:9px;color:#666;
              letter-spacing:1px;text-transform:uppercase;margin-top:4px;}}
.speech-tag{{position:relative;z-index:1;background:rgba(204,0,0,.12);
             border:1px solid rgba(204,0,0,.3);border-radius:4px;
             padding:4px 8px;font-size:9px;color:#bbb;margin-top:6px;
             text-align:center;max-width:110px;line-height:1.4;
             min-height:28px;display:flex;align-items:center;justify-content:center;}}

/* Avatar SVG animations */
@keyframes blink{{
  0%,88%,100%{{transform:scaleY(1)}}
  92%{{transform:scaleY(.07)}}
}}
.eye-l{{animation:blink 3.8s 0s infinite;transform-origin:36px 52px;}}
.eye-r{{animation:blink 3.8s .4s infinite;transform-origin:54px 52px;}}
@keyframes breathe{{
  0%,100%{{transform:translateY(0)}}
  50%{{transform:translateY(-2px)}}
}}
.avatar-wrap svg{{animation:breathe 3s ease-in-out infinite;
                  filter:drop-shadow(0 0 12px rgba(204,0,0,.45));}}
#mouth-open{{display:none;}}
#mouth-closed{{display:block;}}
.speaking #mouth-open{{display:block;}}
.speaking #mouth-closed{{display:none;}}
@keyframes speaking-pulse{{
  0%,100%{{transform:scaleY(1)}}
  50%{{transform:scaleY(0.6)}}
}}
.speaking #mouth-open{{animation:speaking-pulse 0.18s ease-in-out infinite;
                        transform-origin:45px 69px;}}

/* ── Camera panel ────────────────────────────── */
.cam-panel{{
  width:200px;flex-shrink:0;background:#000;
  border:1.5px solid #1E1E1E;border-radius:8px;
  overflow:hidden;position:relative;
}}
#cam-video{{width:200px;height:150px;object-fit:cover;display:block;
            transform:scaleX(-1);}}
#cam-canvas{{position:absolute;top:0;left:0;width:200px;height:150px;
             pointer-events:none;transform:scaleX(-1);}}
.cam-badge{{position:absolute;bottom:5px;left:5px;
            background:rgba(0,0,0,.7);border-radius:3px;
            padding:2px 6px;font-size:9px;color:#ddd;
            display:flex;align-items:center;gap:4px;}}
@keyframes pulse{{0%,100%{{opacity:1}}50%{{opacity:.2}}}}
.dot{{width:6px;height:6px;border-radius:50%;background:#CC0000;
      animation:pulse 1.2s infinite;}}

/* Below camera: permission / status */
.cam-status{{padding:4px 6px;font-size:9px;color:#666;text-align:center;}}

/* ── Score row ───────────────────────────────── */
.scores{{display:flex;gap:6px;}}
.score-card{{
  flex:1;background:#141414;border:1px solid #1E1E1E;border-radius:6px;
  padding:6px 4px;text-align:center;
}}
.score-val{{font-size:20px;font-weight:700;line-height:1;}}
.score-lbl{{font-size:8px;color:#555;text-transform:uppercase;letter-spacing:.5px;margin-top:2px;}}

/* ── Event log ───────────────────────────────── */
.log{{
  flex:1;background:#141414;border:1px solid #1E1E1E;border-radius:6px;
  padding:6px 8px;overflow-y:auto;min-height:60px;max-height:90px;
}}
.log::-webkit-scrollbar{{width:3px;}}
.log::-webkit-scrollbar-thumb{{background:#222;border-radius:2px;}}
.log-row{{font-size:9px;color:#555;line-height:1.7;border-bottom:1px solid #1A1A1A;}}
.log-row.w{{color:#FFA500;}}
.log-row.c{{color:#CC0000;}}

/* ── Status bar ─────────────────────────────── */
.status{{display:flex;align-items:center;gap:5px;font-size:9px;color:#555;}}
.st-dot{{width:6px;height:6px;border-radius:50%;background:#00AA44;flex-shrink:0;}}
.st-dot.w{{background:#FFA500;}}
.st-dot.d{{background:#CC0000;}}
</style>
</head>
<body>
<div class="root">

  <!-- Row 1: avatar + camera -->
  <div class="top">

    <!-- Alex Avatar -->
    <div class="avatar-panel" id="avatarPanel">
      <div class="avatar-bg"></div>
      <div class="avatar-wrap" id="avatarWrap">
        <svg width="88" height="108" viewBox="0 0 90 110" xmlns="http://www.w3.org/2000/svg">
          <!-- shoulders/body -->
          <ellipse cx="45" cy="98" rx="29" ry="13" fill="#1a1a2e"/>
          <!-- neck -->
          <rect x="39" y="66" width="12" height="14" rx="3" fill="#c8a882"/>
          <!-- shirt + collar -->
          <path d="M31 80 Q45 90 59 80 L61 98 Q45 104 29 98 Z" fill="#CC0000"/>
          <!-- tie -->
          <path d="M43 80 L45 96 L47 80 L45 75 Z" fill="#880000"/>
          <!-- head -->
          <ellipse cx="45" cy="51" rx="22" ry="24" fill="#c8a882"/>
          <!-- hair -->
          <ellipse cx="45" cy="30" rx="22" ry="9" fill="#1A0E08"/>
          <rect x="23" y="29" width="5" height="11" rx="2" fill="#1A0E08"/>
          <rect x="62" y="29" width="5" height="11" rx="2" fill="#1A0E08"/>
          <!-- ears -->
          <ellipse cx="23" cy="52" rx="4" ry="5" fill="#bf9b75"/>
          <ellipse cx="67" cy="52" rx="4" ry="5" fill="#bf9b75"/>
          <!-- left eye -->
          <g class="eye-l">
            <ellipse cx="36" cy="51" rx="5" ry="5.5" fill="white"/>
            <ellipse cx="36" cy="52" rx="3" ry="3.5" fill="#1A0E08"/>
            <circle cx="37" cy="50.5" r=".9" fill="white"/>
          </g>
          <!-- right eye -->
          <g class="eye-r">
            <ellipse cx="54" cy="51" rx="5" ry="5.5" fill="white"/>
            <ellipse cx="54" cy="52" rx="3" ry="3.5" fill="#1A0E08"/>
            <circle cx="55" cy="50.5" r=".9" fill="white"/>
          </g>
          <!-- eyebrows -->
          <path d="M30 43 Q36 40 42 43" stroke="#1A0E08" stroke-width="1.8" fill="none" stroke-linecap="round"/>
          <path d="M48 43 Q54 40 60 43" stroke="#1A0E08" stroke-width="1.8" fill="none" stroke-linecap="round"/>
          <!-- nose -->
          <path d="M43 58 Q45 63 47 58" stroke="#aa8860" stroke-width="1.5" fill="none" stroke-linecap="round"/>
          <!-- mouth closed -->
          <path id="mouth-closed" d="M38 68 Q45 72 52 68" stroke="#a0705a" stroke-width="1.8"
                fill="none" stroke-linecap="round"/>
          <!-- mouth open (shown when speaking) -->
          <ellipse id="mouth-open" cx="45" cy="69" rx="7" ry="4" fill="#5a1a0a"/>
        </svg>
      </div>
      <div class="avatar-name">ALEX · Virtual Interviewer</div>
      <div class="speech-tag" id="speechTag">Initialising…</div>
    </div>

    <!-- Camera feed -->
    <div style="display:flex;flex-direction:column;gap:6px;">
      <div class="cam-panel">
        <video id="cam-video" autoplay muted playsinline></video>
        <canvas id="cam-canvas" width="200" height="150"></canvas>
        <div class="cam-badge"><div class="dot"></div>LIVE</div>
      </div>
      <div class="cam-status" id="camStatus">Requesting camera…</div>
    </div>

  </div>

  <!-- Row 2: scores -->
  <div class="scores">
    <div class="score-card">
      <div class="score-val" id="intVal" style="color:#00AA44">100</div>
      <div class="score-lbl">Integrity</div>
    </div>
    <div class="score-card">
      <div class="score-val" id="faceVal" style="color:#0088CC">—</div>
      <div class="score-lbl">Faces</div>
    </div>
    <div class="score-card">
      <div class="score-val" id="evtVal" style="color:#FFA500">0</div>
      <div class="score-lbl">Events</div>
    </div>
  </div>

  <!-- Row 3: event log -->
  <div class="log" id="logPanel">
    <div class="log-row">▶ Proctoring engine ready</div>
  </div>

  <!-- Row 4: status -->
  <div class="status">
    <div class="st-dot" id="stDot"></div>
    <span id="stText">Starting up…</span>
  </div>

</div>

<!-- ═══════════════════════════════════════════════════════════════════════ -->
<script>
// ── Config injected by Python ──────────────────────────────────────────────
const QUIZ_MODE    = {QUIZ_MODE};
const ALLOWED_IDS  = {ALLOWED_IDS};

// ── State ──────────────────────────────────────────────────────────────────
let integrityScore = 100;
let infractions    = 0;
let keyCount       = 0;
let facesNow       = 0;
let lookAwayT      = null;
let lookPenalised  = false;
let camLostT       = null;
let camPenalised   = false;
let infLog         = [];

const LOOK_SECS = 5;
const CAM_SECS  = 10;

// ── DOM refs ───────────────────────────────────────────────────────────────
const intVal    = document.getElementById('intVal');
const faceVal   = document.getElementById('faceVal');
const evtVal    = document.getElementById('evtVal');
const logPanel  = document.getElementById('logPanel');
const stDot     = document.getElementById('stDot');
const stText    = document.getElementById('stText');
const speechTag = document.getElementById('speechTag');
const avatarWrap = document.getElementById('avatarWrap');

// ── Utilities ──────────────────────────────────────────────────────────────
function ts() {{
  return new Date().toLocaleTimeString('en-GB',
    {{hour:'2-digit',minute:'2-digit',second:'2-digit'}});
}}

function addLog(msg, level='n') {{
  const d = document.createElement('div');
  d.className = 'log-row' + (level==='w' ? ' w' : level==='c' ? ' c' : '');
  d.textContent = '['+ts()+'] '+msg;
  logPanel.appendChild(d);
  logPanel.scrollTop = logPanel.scrollHeight;
  if (level!=='n') infLog.push(msg);
}}

function setStatus(txt, level='ok') {{
  stText.textContent = txt;
  stDot.className = 'st-dot' + (level==='w' ? ' w' : level==='d' ? ' d' : '');
}}

function penalty(pts, reason) {{
  infractions++;
  integrityScore = Math.max(0, integrityScore - pts);
  addLog('⚠ '+reason+'  −'+pts+' pts', pts>=20 ? 'c' : 'w');
  updateUI();
  syncToParent();
}}

function updateUI() {{
  intVal.textContent = integrityScore;
  intVal.style.color = integrityScore>=70 ? '#00AA44'
                     : integrityScore>=40 ? '#FFA500' : '#CC0000';
  evtVal.textContent = infractions;
}}

// ── Sync score to parent URL (Python reads on next rerun) ──────────────────
function syncToParent() {{
  try {{
    const u = new URL(window.parent.location.href);
    u.searchParams.set('is', integrityScore);
    u.searchParams.set('ic', infractions);
    window.parent.history.replaceState(null,'',u.toString());
  }} catch(e) {{/* cross-origin in some configs – ignore */}}
}}

// ── Avatar speaking animation ──────────────────────────────────────────────
function setAvatarSpeaking(on, label='') {{
  if (on) avatarWrap.classList.add('speaking');
  else    avatarWrap.classList.remove('speaking');
  speechTag.textContent = label || (on ? 'Speaking…' : 'Ready');
}}

// ── Camera & MediaPipe ─────────────────────────────────────────────────────
const video  = document.getElementById('cam-video');
const canvas = document.getElementById('cam-canvas');
const ctx    = canvas.getContext('2d');

async function startCamera() {{
  try {{
    const stream = await navigator.mediaDevices.getUserMedia({{
      video:{{width:320,height:240,facingMode:'user'}}, audio:true
    }});
    video.srcObject = stream;
    document.getElementById('camStatus').textContent = 'Camera active';
    setStatus('Camera & mic active','ok');
    addLog('✓ Camera and microphone granted');
    initFaceDetection();
    initSpeechRecognition(stream);
    syncToParent();
  }} catch(err) {{
    document.getElementById('camStatus').textContent = 'No camera: '+err.message;
    setStatus('Camera denied','d');
    addLog('✗ Camera denied: '+err.message,'c');
  }}
}}

function initFaceDetection() {{
  try {{
    const fd = new FaceDetection({{
      locateFile: f =>
        'https://cdn.jsdelivr.net/npm/@mediapipe/face_detection@0.4/'+f
    }});
    fd.setOptions({{model:'short', minDetectionConfidence:0.55}});
    fd.onResults(onFaces);
    const cam = new Camera(video, {{
      onFrame: async () => {{ await fd.send({{image:video}}); }},
      width:320, height:240
    }});
    cam.start();
    addLog('✓ MediaPipe face detection active');
  }} catch(e) {{
    addLog('ℹ MediaPipe unavailable – basic monitoring','w');
    startBasicMonitor();
  }}
}}

function onFaces(results) {{
  ctx.clearRect(0,0,200,150);
  const faces = results.detections || [];
  facesNow = faces.length;
  faceVal.textContent = facesNow || '0';

  faces.forEach(d => {{
    const b = d.boundingBox;
    const x = (1 - b.xCenter - b.width/2)*200;
    const y = (b.yCenter - b.height/2)*150;
    ctx.strokeStyle='#00AA44'; ctx.lineWidth=1.5;
    ctx.strokeRect(x,y,b.width*200,b.height*150);
  }});

  const now = Date.now();
  if (facesNow === 0) {{
    if (!lookAwayT)  {{ lookAwayT = now; lookPenalised = false; }}
    if (!camLostT)   {{ camLostT  = now; camPenalised  = false; }}
    if ((now-lookAwayT)/1e3 >= LOOK_SECS && !lookPenalised) {{
      penalty(10,'Looking away > 5 seconds');
      lookPenalised = true;
    }}
    if ((now-camLostT)/1e3 >= CAM_SECS && !camPenalised) {{
      penalty(20,'Camera view unavailable > 10 seconds');
      camPenalised = true;
    }}
  }} else {{
    lookAwayT=null; lookPenalised=false;
    camLostT=null;  camPenalised=false;
  }}
  if (facesNow > 1) penalty(30,'Multiple faces detected in frame');
}}

function startBasicMonitor() {{
  setInterval(() => {{
    faceVal.textContent = video.srcObject ? '1' : '0';
  }}, 2000);
}}

// ── Speech Recognition (voice answers) ────────────────────────────────────
function initSpeechRecognition() {{
  const SR = window.SpeechRecognition || window.webkitSpeechRecognition;
  if (!SR) {{ addLog('ℹ Voice recognition unsupported – use buttons'); return; }}

  const rec = new SR();
  rec.continuous = false; rec.interimResults = false; rec.lang = 'en-US';

  rec.onresult = e => {{
    const t = Array.from(e.results).map(r=>r[0].transcript.trim().toUpperCase()).join(' ');
    addLog('🎤 Heard: "'+t+'"');
    const map = {{
      A:['A','OPTION A','ANSWER A','FIRST','ONE','ALPHA'],
      B:['B','OPTION B','ANSWER B','SECOND','TWO','BRAVO'],
      C:['C','OPTION C','ANSWER C','THIRD','THREE','CHARLIE'],
      D:['D','OPTION D','ANSWER D','FOURTH','FOUR','DELTA'],
    }};
    for (const [letter,phrases] of Object.entries(map)) {{
      if (phrases.some(p=>t.includes(p))) {{
        postAnswer(letter); return;
      }}
    }}
    if (['A','B','C','D'].includes(t.trim())) postAnswer(t.trim());
  }};
  rec.onerror = e => {{ if(e.error!=='no-speech') addLog('⚠ Mic: '+e.error,'w'); }};
  rec.onend   = () => {{ if(QUIZ_MODE) setTimeout(()=>{{ try{{rec.start();}}catch(e){{}} }},600); }};

  if (QUIZ_MODE) {{ try{{rec.start();}}catch(e){{}} }}
  window._rec = rec;
}}

function postAnswer(letter) {{
  addLog('✓ Answer via voice: '+letter);
  try {{
    const u = new URL(window.parent.location.href);
    u.searchParams.set('va',letter);
    window.parent.history.replaceState(null,'',u.toString());
  }} catch(e) {{}}
}}

// ── Event Monitors ─────────────────────────────────────────────────────────
document.addEventListener('keydown', () => {{
  keyCount++;
  penalty(keyCount===1?10:20,
    keyCount===1?'Keyboard activity detected':'Repeated keyboard activity (×'+keyCount+')');
}});

document.addEventListener('visibilitychange', () => {{
  if (document.hidden) penalty(25,'Tab or window switched');
}});

window.addEventListener('blur', () => {{ penalty(25,'Window focus lost'); }});

document.addEventListener('click', e => {{
  if (!QUIZ_MODE) return;
  const id = e.target.id || (e.target.closest('[id]')||{{}}).id || '';
  if (!ALLOWED_IDS.includes(id)) penalty(10,'Click outside answer buttons');
}});

// ── Listen for commands from Streamlit via postMessage ─────────────────────
window.addEventListener('message', e => {{
  const d = e.data; if(!d||!d.cmd) return;
  if (d.cmd==='setAvatarSpeaking') setAvatarSpeaking(d.on, d.label||'');
  if (d.cmd==='setStatus')         setStatus(d.text, d.level||'ok');
  if (d.cmd==='addLog')            addLog(d.text, d.level||'n');
}});

// ── Init ───────────────────────────────────────────────────────────────────
(async ()=>{{
  await startCamera();
  setAvatarSpeaking(false,'Ready for interview');
  updateUI(); syncToParent();
}})();
</script>

<!-- MediaPipe scripts (async so they don't block render) -->
<script async
  src="https://cdn.jsdelivr.net/npm/@mediapipe/face_detection@0.4/face_detection.js"
  crossorigin="anonymous"></script>
<script async
  src="https://cdn.jsdelivr.net/npm/@mediapipe/camera_utils/camera_utils.js"
  crossorigin="anonymous"></script>
</body>
</html>
"""


def build_proctoring_html(quiz_mode: bool = False,
                           allowed_ids: List[str] | None = None) -> str:
    """Fill Python placeholders into the proctoring HTML template."""
    ids = json.dumps(allowed_ids or [])
    return PROCTORING_HTML_TEMPLATE.replace(
        "{QUIZ_MODE}", "true" if quiz_mode else "false"
    ).replace("{ALLOWED_IDS}", ids)


# ─────────────────────────────────────────────────────────────────────────────
# TTS ENGINE  – edge-tts (male neural voice) → gTTS fallback
# ─────────────────────────────────────────────────────────────────────────────
@st.cache_data(show_spinner=False, max_entries=50)
def _generate_tts_bytes(text: str) -> bytes:
    """
    Generate MP3 audio for `text` using Microsoft Edge TTS (en-US-GuyNeural).
    Falls back to gTTS if edge-tts is unavailable.
    Returns raw MP3 bytes, or b'' on total failure.
    """
    # ── Method 1: edge-tts (free, no API key, high-quality male neural voice)
    try:
        import edge_tts  # noqa: PLC0415

        async def _run() -> bytes:
            communicate = edge_tts.Communicate(text, voice="en-US-GuyNeural")
            data = b""
            async for chunk in communicate.stream():
                if chunk["type"] == "audio":
                    data += chunk["data"]
            return data

        loop = asyncio.new_event_loop()
        try:
            return loop.run_until_complete(_run())
        finally:
            loop.close()
    except Exception:
        pass

    # ── Method 2: gTTS (female voice – acceptable fallback)
    try:
        from gtts import gTTS  # noqa: PLC0415
        buf = io.BytesIO()
        gTTS(text=text, lang="en", slow=False).write_to_fp(buf)
        buf.seek(0)
        return buf.read()
    except Exception:
        pass

    return b""


def play_audio_autoplay(audio_bytes: bytes) -> None:
    """
    Inject a hidden <audio autoplay> element into the Streamlit page.
    This plays server-generated speech through the browser's audio stack
    without any user-gesture requirement (because the user has already
    interacted with the Streamlit page by clicking buttons).
    """
    if not audio_bytes:
        return
    b64 = base64.b64encode(audio_bytes).decode()
    st.markdown(
        f'<audio autoplay style="display:none">'
        f'<source src="data:audio/mpeg;base64,{b64}" type="audio/mpeg">'
        f'</audio>',
        unsafe_allow_html=True,
    )


def speak(text: str) -> None:
    """Generate TTS and auto-play it – call once per text you want Alex to say."""
    audio = _generate_tts_bytes(text)
    play_audio_autoplay(audio)


def build_question_tts(q: Dict) -> str:
    """Build the TTS string for a quiz question."""
    opts = "  ".join(f"Option {k}: {v}." for k, v in q["options"].items())
    return f"{q['question']}  {opts}"


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL DATA MANAGER
# ─────────────────────────────────────────────────────────────────────────────
_EXCEL_COLS = [
    "Timestamp", "Employee ID", "Full Name",
    "Questions Attempted", "Correct Answers", "Technical Score (%)",
    "Integrity Score", "Pass / Fail", "Certificate Issued",
    "Duration (mins)", "Infractions Log",
]


def _ensure_excel() -> None:
    if RESULTS_FILE.exists():
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "Assessment Results"
    red  = PatternFill("solid", fgColor="CC0000")
    bold_white = Font(bold=True, color="FFFFFF", size=11)
    center = Alignment(horizontal="center", vertical="center")
    thin   = Side(style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws.row_dimensions[1].height = 28
    for ci, name in enumerate(_EXCEL_COLS, 1):
        c = ws.cell(row=1, column=ci, value=name)
        c.fill = red; c.font = bold_white
        c.alignment = center; c.border = border
        ws.column_dimensions[get_column_letter(ci)].width = max(18, len(name)+4)
    ws.freeze_panes = "A2"
    wb.save(RESULTS_FILE)


def save_result(
    *,
    employee_id: str, full_name: str,
    questions_attempted: int, correct_answers: int,
    integrity_score: int, duration_seconds: float,
    infractions_log: List[str],
) -> Dict[str, Any]:
    _ensure_excel()
    tech = round(correct_answers / max(questions_attempted, 1) * 100, 1)
    passed = tech >= PASS_THRESHOLD
    row = [
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        employee_id, full_name,
        questions_attempted, correct_answers, tech,
        integrity_score,
        "PASS" if passed else "FAIL",
        "Yes" if passed else "No",
        round(duration_seconds / 60, 1),
        " | ".join(infractions_log) if infractions_log else "None",
    ]
    wb = load_workbook(RESULTS_FILE)
    ws = wb.active
    nr = ws.max_row + 1
    thin   = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    alt    = PatternFill("solid", fgColor="FFF5F5") if nr % 2 == 0 else None
    for ci, val in enumerate(row, 1):
        c = ws.cell(row=nr, column=ci, value=val)
        c.border = border
        c.alignment = Alignment(
            horizontal="center",
            wrap_text=(ci == len(_EXCEL_COLS)),
        )
        if alt:
            c.fill = alt
    status_cell = ws.cell(row=nr, column=8)
    status_cell.font = Font(bold=True, color="006600" if passed else "CC0000")
    wb.save(RESULTS_FILE)
    return {"technical_score": tech, "passed": passed}


# ─────────────────────────────────────────────────────────────────────────────
# CERTIFICATE GENERATOR  (ReportLab)
# ─────────────────────────────────────────────────────────────────────────────
def generate_certificate(
    full_name: str,
    employee_id: str,
    technical_score: float,
    integrity_score: int,
) -> bytes:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.pdfgen import canvas as rl_canvas

    W, H = landscape(A4)
    buf  = io.BytesIO()
    c    = rl_canvas.Canvas(buf, pagesize=landscape(A4))

    # Background
    c.setFillColor(colors.HexColor("#0F0F0F"))
    c.rect(0, 0, W, H, fill=1, stroke=0)

    # Red side bars
    c.setFillColor(colors.HexColor("#CC0000"))
    c.rect(0, 0, 16, H, fill=1, stroke=0)
    c.rect(W-16, 0, 16, H, fill=1, stroke=0)

    # Subtle grid
    c.setStrokeColor(colors.HexColor("#181818"))
    c.setLineWidth(0.3)
    for x in range(30, int(W), 18):
        c.line(x, 0, x, H)
    for y in range(0, int(H), 18):
        c.line(30, y, W-30, y)

    # Inner card
    m = 26
    c.setFillColor(colors.HexColor("#141414"))
    c.setStrokeColor(colors.HexColor("#CC0000"))
    c.setLineWidth(1.8)
    c.roundRect(m+16, m, W-2*m-32, H-2*m, 8, fill=1, stroke=1)

    # Header strip
    c.setFillColor(colors.HexColor("#CC0000"))
    c.rect(m+16, H-m-54, W-2*m-32, 54, fill=1, stroke=0)
    c.setFillColor(colors.white)
    c.setFont("Helvetica-Bold", 20)
    c.drawString(m+32, H-m-36, "LENOVO")
    c.setFont("Helvetica", 9)
    c.setFillColor(colors.HexColor("#FFCCCC"))
    c.drawRightString(W-m-32, H-m-36, "ELITE ENGINEER CERTIFICATION PROGRAM")

    # Title
    c.setFillColor(colors.white)
    c.setFont("Helvetica-Bold", 34)
    c.drawCentredString(W/2, H-m-108, "Certificate of Achievement")
    c.setStrokeColor(colors.HexColor("#CC0000"))
    c.setLineWidth(1.4)
    c.line(W/2-120, H-m-120, W/2+120, H-m-120)

    # Recipient
    c.setFillColor(colors.HexColor("#AAAAAA"))
    c.setFont("Helvetica", 11)
    c.drawCentredString(W/2, H-m-150, "This is to certify that")
    c.setFillColor(colors.white)
    c.setFont("Helvetica-BoldOblique", 28)
    c.drawCentredString(W/2, H-m-190, full_name)
    nw = c.stringWidth(full_name, "Helvetica-BoldOblique", 28)
    c.setStrokeColor(colors.HexColor("#CC0000"))
    c.setLineWidth(0.8)
    c.line(W/2-nw/2, H-m-196, W/2+nw/2, H-m-196)

    c.setFillColor(colors.HexColor("#AAAAAA"))
    c.setFont("Helvetica", 11)
    c.drawCentredString(W/2, H-m-220,
        "has successfully completed the Lenovo Advanced PC Hardware Troubleshooting Assessment")
    c.drawCentredString(W/2, H-m-238, "and has been awarded the designation of")
    c.setFillColor(colors.HexColor("#FF4444"))
    c.setFont("Helvetica-Bold", 17)
    c.drawCentredString(W/2, H-m-264, "LENOVO ELITE ENGINEER")

    # Score boxes
    for cx, lbl, val, col in [
        (W/2-150, "Technical Score", f"{technical_score}%", "#00AA44"),
        (W/2+150, "Integrity Score",  f"{integrity_score}/100", "#0088CC"),
    ]:
        c.setFillColor(colors.HexColor("#1E1E1E"))
        c.setStrokeColor(colors.HexColor(col))
        c.setLineWidth(1.4)
        c.roundRect(cx-65, m+44, 130, 52, 6, fill=1, stroke=1)
        c.setFillColor(colors.HexColor(col))
        c.setFont("Helvetica-Bold", 19)
        c.drawCentredString(cx, m+76, val)
        c.setFillColor(colors.HexColor("#888888"))
        c.setFont("Helvetica", 8)
        c.drawCentredString(cx, m+58, lbl)

    # Date + ID
    c.setFillColor(colors.HexColor("#555555"))
    c.setFont("Helvetica", 8)
    cert_date = datetime.now().strftime("%B %d, %Y")
    c.drawCentredString(W/2, m+30,
        f"Employee ID: {employee_id}   |   Date of Issue: {cert_date}")

    # Footer
    c.setFont("Helvetica", 7)
    c.drawCentredString(W/2, m+10,
        "This certificate is electronically generated and is valid for verification through the Lenovo Training Portal.")

    c.save()
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
# SESSION STATE
# ─────────────────────────────────────────────────────────────────────────────
def init_state() -> None:
    defaults: Dict[str, Any] = {
        "phase":          "registration",
        "name":           "",
        "employee_id":    "",
        "q_index":        0,
        "answers":        {},
        "start_time":     None,
        "integrity_score": 100,
        "infractions_log": [],
        "intro_played":   False,
        "q_audio_played": -1,
        "result_saved":   False,
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v


init_state()


# ─────────────────────────────────────────────────────────────────────────────
# GLOBAL CSS
# ─────────────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
#MainMenu,footer,header,.stDeployButton{visibility:hidden}
.stApp,[data-testid="stAppViewContainer"]{background:#0F0F0F!important}
[data-testid="stVerticalBlock"]{gap:.4rem}

/* cards */
.card{background:#141414;border:1.5px solid #1E1E1E;border-radius:10px;padding:22px 26px}

/* top header bar */
.app-header{
  background:linear-gradient(135deg,#CC0000 0%,#880000 100%);
  padding:13px 26px;border-radius:10px;
  display:flex;align-items:center;justify-content:space-between;
  margin-bottom:16px;
}
.app-header h1{font-size:1.2rem;font-weight:700;color:#fff;margin:0;letter-spacing:.4px}
.app-header .badge{
  font-size:.68rem;background:rgba(255,255,255,.18);color:#fff;
  padding:3px 10px;border-radius:20px;text-transform:uppercase;letter-spacing:.5px
}

/* inputs */
.stTextInput>div>div>input{
  background:#1A1A1A!important;border:1.5px solid #2A2A2A!important;
  color:#E0E0E0!important;border-radius:8px!important;padding:10px 14px!important
}
.stTextInput>div>div>input:focus{
  border-color:#CC0000!important;box-shadow:0 0 0 3px rgba(204,0,0,.15)!important
}
label[data-testid="stWidgetLabel"]{color:#888!important;font-size:.82rem!important}

/* buttons */
.stButton>button{
  background:linear-gradient(135deg,#CC0000,#880000)!important;
  color:#fff!important;border:none!important;border-radius:8px!important;
  padding:11px 26px!important;font-weight:600!important;font-size:.93rem!important;
  width:100%;transition:opacity .15s
}
.stButton>button:hover{opacity:.86}
.stButton>button:disabled{opacity:.35!important;cursor:not-allowed!important}

/* progress bar */
.prog-wrap{background:#1A1A1A;border-radius:999px;height:5px;margin-bottom:18px;overflow:hidden}
.prog-fill{height:100%;background:linear-gradient(90deg,#CC0000,#FF4444);
           border-radius:999px;transition:width .4s ease}

/* option buttons */
div[data-testid="stHorizontalBlock"] .stButton>button,
.opt-btn .stButton>button{text-align:left!important}
</style>
""", unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────────────────────
# SHARED HELPERS
# ─────────────────────────────────────────────────────────────────────────────
def render_header(subtitle: str = "") -> None:
    st.markdown(
        f'<div class="app-header"><h1>🔧 LENOVO ELITE ENGINEER CERTIFICATION</h1>'
        f'<span class="badge">{subtitle}</span></div>',
        unsafe_allow_html=True,
    )


def render_proctoring(quiz_mode: bool = False,
                      height: int = 400,
                      allowed_ids: List[str] | None = None) -> None:
    """Render the inline proctoring + avatar component."""
    html = build_proctoring_html(quiz_mode=quiz_mode, allowed_ids=allowed_ids)
    components.html(html, height=height, scrolling=False)


def sync_integrity_from_params() -> None:
    """Pull integrity score / infractions from JS-written URL query params."""
    try:
        params = st.query_params
        if "is" in params:
            st.session_state.integrity_score = max(
                0, min(100, int(params["is"]))
            )
    except Exception:
        pass


def card(content_fn, **kwargs):
    """Wrap content in a dark card div."""
    st.markdown('<div class="card">', unsafe_allow_html=True)
    content_fn(**kwargs)
    st.markdown("</div>", unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────────────────────
# PHASE 1 – REGISTRATION
# ─────────────────────────────────────────────────────────────────────────────
def phase_registration() -> None:
    render_header("Step 1 of 4 · Registration")

    left, right = st.columns([1.1, 0.9], gap="large")

    with left:
        st.markdown('<div class="card">', unsafe_allow_html=True)
        st.markdown("### Candidate Registration")
        st.markdown(
            '<p style="color:#666;font-size:.83rem;margin-bottom:18px">'
            "Enter your details to begin the proctored certification assessment.</p>",
            unsafe_allow_html=True,
        )
        name   = st.text_input("Full Name",    placeholder="e.g. Ramesh Kumar",    key="ri_name")
        emp_id = st.text_input("Employee ID",  placeholder="e.g. LEN-2025-0042",  key="ri_id")
        st.markdown(
            '<div style="background:#1A1A1A;border:1px solid #2A2A2A;border-radius:8px;'
            'padding:12px 14px;margin:14px 0;font-size:.8rem;color:#666;">'
            '<strong style="color:#CC0000">ℹ Before you begin</strong><br/>'
            '• Camera + microphone will be activated<br/>'
            '• Alex (AI interviewer) will speak instructions aloud<br/>'
            '• 11 technical questions · passing score: <strong style="color:#ddd">90 %</strong>'
            '</div>',
            unsafe_allow_html=True,
        )
        if st.button("Proceed to Assessment →", key="btn_reg"):
            if not name.strip():
                st.error("Please enter your full name.")
            elif not emp_id.strip():
                st.error("Please enter your Employee ID.")
            else:
                st.session_state.name        = name.strip()
                st.session_state.employee_id = emp_id.strip()
                st.session_state.phase       = "onboarding"
                st.rerun()
        st.markdown("</div>", unsafe_allow_html=True)

    with right:
        st.markdown(
            '<div class="card" style="text-align:center">'
            '<div style="font-size:3rem;margin-bottom:10px">🎓</div>'
            '<h3 style="color:#CC0000;margin-bottom:8px">Elite Engineer Program</h3>'
            '<p style="color:#666;font-size:.82rem;line-height:1.65">'
            "Validates advanced competency in Lenovo ThinkPad hardware troubleshooting — "
            "No-Power diagnostics, POST analysis, display fault isolation, and Crisis BIOS recovery."
            "</p>"
            '<hr style="border:1px solid #1E1E1E;margin:14px 0"/>'
            '<div style="display:flex;justify-content:space-around">'
            '<div><div style="font-size:1.4rem;font-weight:700;color:#E0E0E0">11</div>'
            '<div style="font-size:.68rem;color:#555;text-transform:uppercase">Questions</div></div>'
            '<div><div style="font-size:1.4rem;font-weight:700;color:#E0E0E0">90 %</div>'
            '<div style="font-size:.68rem;color:#555;text-transform:uppercase">Pass Mark</div></div>'
            '<div><div style="font-size:1.4rem;font-weight:700;color:#E0E0E0">~15 m</div>'
            '<div style="font-size:.68rem;color:#555;text-transform:uppercase">Duration</div></div>'
            "</div></div>",
            unsafe_allow_html=True,
        )


# ─────────────────────────────────────────────────────────────────────────────
# PHASE 2 – ONBOARDING
# ─────────────────────────────────────────────────────────────────────────────
def phase_onboarding() -> None:
    render_header("Step 2 of 4 · Briefing & Permissions")
    sync_integrity_from_params()

    # ── Auto-play Alex intro ONCE (server-side edge-tts, no gesture needed) ──
    if not st.session_state.intro_played:
        with st.spinner("Alex is preparing your briefing…"):
            speak(ALEX_INTRO)
        st.session_state.intro_played = True

    left, right = st.columns([1.4, 1], gap="large")

    with left:
        st.markdown(
            f'<div class="card" style="margin-bottom:12px">'
            f'<p style="color:#666;font-size:.78rem;margin:0">Candidate</p>'
            f'<h2 style="margin:4px 0 0;font-size:1.35rem">Hello, {st.session_state.name} 👋</h2>'
            f'<p style="color:#555;font-size:.78rem;margin:3px 0 0">ID: {st.session_state.employee_id}</p>'
            f'</div>',
            unsafe_allow_html=True,
        )

        # Proctoring + avatar component
        render_proctoring(quiz_mode=False, height=400)

        # Volume note
        st.markdown(
            '<p style="font-size:.75rem;color:#444;margin-top:6px;text-align:center">'
            '🔊 Alex\'s introduction is playing via your speakers. '
            'Ensure volume is turned up.</p>',
            unsafe_allow_html=True,
        )

    with right:
        st.markdown('<div class="card">', unsafe_allow_html=True)
        st.markdown("#### 📋 Interview Rules")
        st.markdown(
            '<ul style="color:#666;font-size:.82rem;line-height:2.1;padding-left:16px">'
            "<li>Alex reads each question aloud automatically</li>"
            "<li>Answer <em>verbally</em> or click an option button</li>"
            "<li><strong style='color:#CC0000'>No keyboard input</strong> is permitted</li>"
            "<li>Keep your face clearly visible at all times</li>"
            "<li>No other people in the camera frame</li>"
            "<li>Do not switch tabs or minimise the window</li>"
            "</ul>",
            unsafe_allow_html=True,
        )
        st.markdown(
            '<div style="background:rgba(204,0,0,.07);border:1px solid rgba(204,0,0,.25);'
            'border-radius:8px;padding:13px;margin-top:12px">'
            '<strong style="color:#CC0000;font-size:.83rem">⚡ Integrity Penalties</strong>'
            '<table style="width:100%;margin-top:8px;font-size:.77rem;'
            'color:#666;border-collapse:collapse">'
            "<tr><td>Keyboard activity</td><td style='text-align:right;color:#FF6666'>−10 pts</td></tr>"
            "<tr><td>Tab / window switch</td><td style='text-align:right;color:#FF6666'>−25 pts</td></tr>"
            "<tr><td>Look away &gt; 5 s</td><td style='text-align:right;color:#FF6666'>−10 pts</td></tr>"
            "<tr><td>Multiple faces</td><td style='text-align:right;color:#FF6666'>−30 pts</td></tr>"
            "<tr><td>Phone detected</td><td style='text-align:right;color:#FF6666'>−30 pts</td></tr>"
            "<tr><td>Camera lost &gt; 10 s</td><td style='text-align:right;color:#FF6666'>−20 pts</td></tr>"
            "</table></div>",
            unsafe_allow_html=True,
        )
        st.markdown("</div>", unsafe_allow_html=True)

        with st.expander("📄 Read Alex's script", expanded=False):
            st.markdown(
                f'<p style="color:#555;font-size:.77rem;line-height:1.7">{ALEX_INTRO}</p>',
                unsafe_allow_html=True,
            )

        st.markdown("<br/>", unsafe_allow_html=True)
        if st.button("✅  I'm Ready – Begin Assessment", key="btn_ready"):
            st.session_state.phase      = "quiz"
            st.session_state.q_index   = 0
            st.session_state.answers   = {}
            st.session_state.start_time = time.time()
            st.session_state.q_audio_played = -1
            st.rerun()


# ─────────────────────────────────────────────────────────────────────────────
# PHASE 3 – QUIZ
# ─────────────────────────────────────────────────────────────────────────────
def phase_quiz() -> None:
    sync_integrity_from_params()

    qi = st.session_state.q_index
    nq = len(QUESTIONS)

    if qi >= nq:
        st.session_state.phase = "results"
        st.rerun()
        return

    q = QUESTIONS[qi]
    render_header(f"Step 3 of 4 · Question {qi+1} of {nq}")

    # Progress bar
    pct = int(qi / nq * 100)
    st.markdown(
        f'<div class="prog-wrap"><div class="prog-fill" style="width:{pct}%"></div></div>',
        unsafe_allow_html=True,
    )

    qleft, qright = st.columns([1.5, 1], gap="large")

    # ── Question + answers ──────────────────────────────────────────────────
    with qleft:
        st.markdown('<div class="card">', unsafe_allow_html=True)

        if q.get("topic"):
            st.markdown(
                f'<span style="font-size:.7rem;color:#CC0000;text-transform:uppercase;'
                f'letter-spacing:.8px;font-weight:600">{q["topic"]}</span>',
                unsafe_allow_html=True,
            )

        st.markdown(
            f'<h3 style="margin:10px 0 20px;font-size:1.02rem;line-height:1.55;color:#E0E0E0">'
            f'Q{qi+1}. {q["question"]}</h3>',
            unsafe_allow_html=True,
        )

        # Auto-play question TTS once when a new question loads
        if st.session_state.q_audio_played != qi:
            speak(build_question_tts(q))
            st.session_state.q_audio_played = qi

        # Answer buttons
        selected = st.session_state.answers.get(qi)
        for letter, text in q["options"].items():
            if selected:
                # Disable all options after answering
                c = "✅" if letter == q["answer"] else ("❌" if letter == selected else "  ")
                colour = "#006600" if letter == q["answer"] else \
                         ("#660000" if letter == selected else "#2A2A2A")
                st.markdown(
                    f'<div style="background:#141414;border:1.5px solid {colour};'
                    f'border-radius:8px;padding:13px 16px;margin-bottom:9px;'
                    f'font-size:.9rem;color:#888">'
                    f'{c} &nbsp;<strong style="color:#aaa">{letter}.</strong> {text}'
                    f'</div>',
                    unsafe_allow_html=True,
                )
            else:
                btn_id = f"opt_{letter}_{qi}"
                if st.button(
                    f"  {letter}.  {text}",
                    key=btn_id,
                ):
                    st.session_state.answers[qi] = letter
                    st.rerun()

        # Explanation
        if selected and q.get("explanation"):
            ok     = selected == q["answer"]
            colour = "#00AA44" if ok else "#CC0000"
            icon   = "✅" if ok else "❌"
            label  = "Correct!" if ok else f"Incorrect – correct answer: {q['answer']}"
            st.markdown(
                f'<div style="background:rgba(255,255,255,.02);border:1px solid {colour}44;'
                f'border-radius:8px;padding:11px 14px;margin-top:14px;font-size:.82rem">'
                f'<strong style="color:{colour}">{icon} {label}</strong><br/>'
                f'<span style="color:#666">{q["explanation"]}</span>'
                f'</div>',
                unsafe_allow_html=True,
            )

        st.markdown("</div>", unsafe_allow_html=True)

        if selected:
            st.markdown("<br/>", unsafe_allow_html=True)
            lbl = "Next Question →" if qi < nq - 1 else "Finish Assessment ✓"
            if st.button(lbl, key=f"btn_next_{qi}"):
                st.session_state.q_index += 1
                st.session_state.q_audio_played = -1   # reset so next Q auto-plays
                st.rerun()

    # ── Proctoring panel ────────────────────────────────────────────────────
    with qright:
        allowed = [f"opt_{l}_{qi}" for l in q["options"]]
        render_proctoring(quiz_mode=True, height=400, allowed_ids=allowed)

        # Score card below component
        iscore = st.session_state.integrity_score
        icol   = "#00AA44" if iscore >= 70 else "#FFA500" if iscore >= 40 else "#CC0000"
        st.markdown(
            f'<div class="card" style="text-align:center;padding:14px;margin-top:8px">'
            f'<div style="font-size:.68rem;color:#555;text-transform:uppercase;'
            f'letter-spacing:.8px;margin-bottom:4px">Integrity Score</div>'
            f'<div style="font-size:2rem;font-weight:700;color:{icol}">'
            f'{iscore}<span style="font-size:.9rem;color:#444">/100</span></div>'
            f'</div>',
            unsafe_allow_html=True,
        )


# ─────────────────────────────────────────────────────────────────────────────
# PHASE 4 – RESULTS
# ─────────────────────────────────────────────────────────────────────────────
def phase_results() -> None:
    sync_integrity_from_params()
    render_header("Step 4 of 4 · Results & Certification")

    nq      = len(QUESTIONS)
    correct = sum(1 for i, q in enumerate(QUESTIONS)
                  if st.session_state.answers.get(i) == q["answer"])
    tech    = round(correct / max(nq, 1) * 100, 1)
    iscore  = st.session_state.integrity_score
    passed  = tech >= PASS_THRESHOLD
    dur     = time.time() - (st.session_state.start_time or time.time())

    # Save to Excel exactly once
    if not st.session_state.result_saved:
        try:
            save_result(
                employee_id=st.session_state.employee_id,
                full_name=st.session_state.name,
                questions_attempted=nq,
                correct_answers=correct,
                integrity_score=iscore,
                duration_seconds=dur,
                infractions_log=st.session_state.infractions_log,
            )
        except Exception as e:
            st.warning(f"Excel logging failed: {e}")
        st.session_state.result_saved = True

    # Play result audio
    result_msg = (
        f"Congratulations {st.session_state.name}! "
        f"You have passed the Lenovo Elite Engineer Certification with a technical score of "
        f"{tech} percent and an integrity score of {iscore} out of 100. "
        f"Your certificate is ready to download."
        if passed else
        f"Assessment complete, {st.session_state.name}. "
        f"Your technical score was {tech} percent. "
        f"You need 90 percent or higher to earn the certificate. Keep studying and try again!"
    )
    speak(result_msg)

    rleft, rright = st.columns([1.2, 1], gap="large")

    with rleft:
        # Banner
        if passed:
            st.markdown(
                '<div style="background:linear-gradient(135deg,#081808,#051305);'
                'border:1.5px solid #00AA44;border-radius:10px;'
                'padding:24px;text-align:center;margin-bottom:14px">'
                '<div style="font-size:2.6rem;margin-bottom:8px">🏆</div>'
                '<h2 style="color:#00AA44;margin:0 0 6px">Congratulations!</h2>'
                '<p style="color:#666;font-size:.86rem;margin:0">'
                'You have passed the Lenovo Elite Engineer Certification.</p>'
                '</div>',
                unsafe_allow_html=True,
            )
        else:
            st.markdown(
                '<div style="background:#100A0A;border:1.5px solid #CC0000;'
                'border-radius:10px;padding:24px;text-align:center;margin-bottom:14px">'
                '<div style="font-size:2.6rem;margin-bottom:8px">📋</div>'
                '<h2 style="color:#CC0000;margin:0 0 6px">Assessment Complete</h2>'
                '<p style="color:#666;font-size:.86rem;margin:0">'
                'A score of 90 % or higher is required to earn the certificate.</p>'
                '</div>',
                unsafe_allow_html=True,
            )

        # Score row
        s1, s2, s3 = st.columns(3)
        for col, label, val, colour in [
            (s1, "Technical",  f"{tech}%", "#00AA44" if passed else "#CC0000"),
            (s2, "Integrity",  f"{iscore}", "#00AA44" if iscore>=70 else "#FFA500" if iscore>=40 else "#CC0000"),
            (s3, "Correct",    f"{correct}/{nq}", "#E0E0E0"),
        ]:
            with col:
                st.markdown(
                    f'<div class="card" style="text-align:center;padding:14px">'
                    f'<div style="font-size:1.7rem;font-weight:700;color:{colour}">{val}</div>'
                    f'<div style="font-size:.67rem;color:#555;text-transform:uppercase;margin-top:3px">{label}</div>'
                    f'</div>',
                    unsafe_allow_html=True,
                )

        # Question breakdown
        st.markdown('<div class="card" style="margin-top:12px">', unsafe_allow_html=True)
        st.markdown('<p style="font-weight:600;font-size:.9rem;margin-bottom:10px">Question Breakdown</p>',
                    unsafe_allow_html=True)
        for i, q in enumerate(QUESTIONS):
            ua  = st.session_state.answers.get(i, "—")
            ok  = ua == q["answer"]
            col = "#00AA44" if ok else "#CC0000"
            ico = "✅" if ok else "❌"
            preview = q["question"][:55] + "…"
            st.markdown(
                f'<div style="display:flex;justify-content:space-between;'
                f'padding:6px 0;border-bottom:1px solid #1A1A1A;'
                f'font-size:.79rem;color:#666">'
                f'<span>{ico} Q{i+1}: {preview}</span>'
                f'<span style="color:{col};white-space:nowrap;margin-left:8px">'
                f'{ua} / {q["answer"]}</span></div>',
                unsafe_allow_html=True,
            )
        st.markdown("</div>", unsafe_allow_html=True)

    with rright:
        if passed:
            st.markdown('<div class="card" style="text-align:center">', unsafe_allow_html=True)
            st.markdown(
                '<div style="font-size:2rem;margin-bottom:8px">🎓</div>'
                '<h4 style="margin-bottom:8px">Your Certificate</h4>'
                '<p style="color:#666;font-size:.8rem">Download your official PDF certificate below.</p>',
                unsafe_allow_html=True,
            )
            try:
                cert_bytes = generate_certificate(
                    full_name=st.session_state.name,
                    employee_id=st.session_state.employee_id,
                    technical_score=tech,
                    integrity_score=iscore,
                )
                b64   = base64.b64encode(cert_bytes).decode()
                fname = f"Lenovo_Elite_Engineer_{st.session_state.name.replace(' ','_')}.pdf"
                st.markdown(
                    f'<a href="data:application/pdf;base64,{b64}" download="{fname}"'
                    f' style="display:inline-block;background:linear-gradient(135deg,#CC0000,#880000);'
                    f'color:#fff;padding:12px 28px;border-radius:8px;font-weight:600;'
                    f'text-decoration:none;font-size:.93rem;margin-top:10px">'
                    f'⬇ Download Certificate (PDF)</a>',
                    unsafe_allow_html=True,
                )
            except Exception as e:
                st.error(f"Certificate error: {e}")
                st.info("Ensure `reportlab` is installed: `pip install reportlab`")
            st.markdown("</div>", unsafe_allow_html=True)
        else:
            st.markdown(
                '<div class="card" style="text-align:center">'
                '<div style="font-size:1.8rem;margin-bottom:10px">📖</div>'
                '<h4 style="margin-bottom:8px">Keep Studying</h4>'
                '<p style="color:#666;font-size:.8rem;line-height:1.6">'
                'Review the topics below, then retake the assessment '
                'to earn your Elite Engineer certificate.</p>',
                unsafe_allow_html=True,
            )
            for q in QUESTIONS:
                if q.get("topic"):
                    ua = st.session_state.answers.get(QUESTIONS.index(q))
                    if ua != q["answer"]:
                        st.markdown(
                            f'<div style="font-size:.78rem;color:#666;padding:3px 0">📌 {q["topic"]}</div>',
                            unsafe_allow_html=True,
                        )
            st.markdown("</div>", unsafe_allow_html=True)

        # Excel download
        if RESULTS_FILE.exists():
            st.markdown("<br/>", unsafe_allow_html=True)
            with open(RESULTS_FILE, "rb") as f:
                st.download_button(
                    "📊 Download Results Log (Excel)",
                    data=f.read(),
                    file_name="assessment_results.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

        st.markdown("<br/>", unsafe_allow_html=True)
        if st.button("🔄 Start New Assessment", key="btn_restart"):
            # Clear all state
            for k in list(st.session_state.keys()):
                del st.session_state[k]
            # Clear score query params
            try:
                st.query_params.clear()
            except Exception:
                pass
            st.rerun()


# ─────────────────────────────────────────────────────────────────────────────
# ROUTER
# ─────────────────────────────────────────────────────────────────────────────
phase = st.session_state.get("phase", "registration")

if   phase == "registration": phase_registration()
elif phase == "onboarding":   phase_onboarding()
elif phase == "quiz":         phase_quiz()
elif phase == "results":      phase_results()
else:
    st.session_state.phase = "registration"
    st.rerun()
